import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ═══════════════════════════ COLORS ═══════════════════════════
ORANGE_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
BLUE_FILL   = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
RED_FILL    = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
PURPLE_FILL = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')
GREEN_FILL  = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
HEAD_FILL = PatternFill(start_color='263238', end_color='263238', fill_type='solid')
WHITE_FONT = Font(color='FFFFFF', bold=True, size=10)
BOLD = Font(bold=True, size=10)
BOLD_RED = Font(bold=True, size=9, color='C62828')
BOLD_GREEN = Font(bold=True, size=9, color='2E7D32')
BOLD_BLUE = Font(bold=True, size=9, color='1565C0')
BOLD_ORANGE = Font(bold=True, size=9, color='E65100')
BOLD_PURPLE = Font(bold=True, size=9, color='7B1FA2')
NORMAL = Font(size=9)
SMALL = Font(size=8, color='757575')
thin_border = Border(left=Side('thin','BDBDBD'),right=Side('thin','BDBDBD'),
    top=Side('thin','BDBDBD'),bottom=Side('thin','BDBDBD'))
AL = Alignment(vertical='center', wrap_text=True)
AL_C = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ═══════════════════════════ SHEET 1 ═══════════════════════════
ws = wb.active
ws.title = "五节点建仓地图"

# Title
ws.merge_cells('A1:O1'); ws['A1'] = '⚔️ 向心坍缩 · 建仓日历 v2.0 · 国内投资者作战地图 · 2026.07.31'
ws['A1'].font = Font(bold=True, size=16, color='263238'); ws['A1'].alignment = Alignment(horizontal='center'); ws.row_dimensions[1].height = 35
ws.merge_cells('A2:O2'); ws['A2'] = 'GOR=47.9 | WTI=$85.56 | 10Y=4.67% | DXY=100.12 | VIX=17.09 | 当前: 节点一·预热建仓期 | 仓位目标: 37%→50%'
ws['A2'].font = SMALL; ws['A2'].alignment = Alignment(horizontal='center'); ws.row_dimensions[2].height = 20

# Headers
h = ['节点','时间段','核心事件','资产','A股标的/代码','买点(人民币)','仓位变动','止损/止盈','观测信号','⚠ 仓位把控']
for c, v in enumerate(h, 1):
    cell = ws.cell(row=3, column=c, value=v)
    cell.font = WHITE_FONT; cell.fill = HEAD_FILL; cell.alignment = AL_C; cell.border = thin_border
ws.row_dimensions[3].height = 25

# Data — 5 nodes × 7 assets
nodes = [
    # Node 1: Warm-up (orange)
    ('节点一\n预热建仓','7月末–8月','布伦特95-110\n美债发行预热\n分批建仓·不追高', ORANGE_FILL, BOLD_ORANGE, [
        ('🛢️ 原油','华宝油气LOF/中石油','布伦特95-100回踩','3%→5%','布伦特<88减50%','霍尔木兹日通行<10艘','总仓37→50%'),
        ('🪨 煤炭','中国神华601088/陕西煤业','神华30-32元','5%→8%','神华<28元减仓','纽卡斯尔煤价周涨',''),
        ('⚙️ 铝','云铝000807/天山铝业','沪铝19500-20000','4%→6%','沪铝<18500减仓','LME铝库存周降',''),
        ('🔶 铜','紫金矿业601899','紫金19-20元','3%→5%','沪铜<70000减仓','COMEX铜库存',''),
        ('💎 黄金','黄金ETF 518880','现货4000-4200$\n518880约9.0-9.5','10%→12%','现货<3850减20%','央行月度购金量',''),
        ('🥈 白银','盛达资源000603','现货55-58$\n盛达12-13元','2%→3%','现货<50离场','金银比>80',''),
        ('📜 国债','国债ETF 511260','10Y 3.2-3.4%配置','10%→12%','收益率<3.0%转','CPI+央行政策',''),
    ]),
    # Node 2: Acceleration (blue)
    ('节点二\n加仓期','8月–9月中','布伦特100-110确认\n加码主线·趋势确认', BLUE_FILL, BOLD_BLUE, [
        ('🛢️ 原油','华宝油气','5%→7%','布伦特100-110确认','通航恢复→减','霍尔木兹<10艘','总仓50→61%'),
        ('🪨 煤炭','中国神华','8%→10%','神华32-35持有','保供限价→减','发改委政策',''),
        ('⚙️ 铝','云铝/天山铝业','6%→8%','中东减产确认','沪铝>21000不追','中东铝厂公告',''),
        ('🔶 铜','紫金/洛阳钼业','5%→7%','紫金20-21元','DXY>104暂缓','美债投标倍数',''),
        ('💎 黄金','518880','12%→13%','4000以下额外加','央行停购→减2%','欧/日央行购金',''),
        ('🥈 白银','盛达资源','3%→4%','跟随黄金','工业需求弱→减','光伏装机数据',''),
        ('📜 国债','511260','维持12%','高票息持有','10Y>3.5%加','收益率走势',''),
    ]),
    # Node 3: Profit-taking (red)
    ('节点三\n兑现+对冲','9月中–10月初','布伦特120-150极值\n兑现利润·不追高', RED_FILL, BOLD_RED, [
        ('🛢️ 原油 ⚠️','华宝油气','7%→3-4%','布伦特120+减50%','不追高！','美债投标<2.0','总仓61→53%'),
        ('🪨 煤炭','中国神华','10%→7%','神华35+兑现30%','限价风险↑','煤价政策',''),
        ('⚙️ 铝','云铝','8%→5%','沪铝22000+分批','兑现40%','LME库存',''),
        ('🔶 铜','紫金','7%→5%','紫金21+兑现30%','留底仓','COMEX头寸',''),
        ('💎 黄金','518880','13%→12%','持仓不动！','3800以下加','央行购金公告',''),
        ('🥈 白银','盛达','4%','不追高','波动放大','',''),
        ('📜 国债','511260','12%→14%','避险+票息','速冻期最稳','离岸美元基差',''),
    ]),
    # Node 4: Freeze Buy (purple, star)
    ('节点四⭐\n速冻买点','10月–11月','美元荒·全资产回调\n⭐最佳二次布局', PURPLE_FILL, BOLD_PURPLE, [
        ('💎 黄金 ⭐⭐⭐','518880','12%→18%','518880回踩8.2-8.8元','现货<3600减10%','EUR/USD xccy走阔','总仓53→69%'),
        ('⚙️ 铝 ⭐⭐⭐','云铝','5%→12%','沪铝18500-19500','沪铝<17500','中东铝厂停产','⚠ 不突破红线①'),
        ('🔶 铜 ⭐⭐','紫金','5%→10%','沪铜68000-72000','沪铜<65000','COMEX净头寸转负','单标≤5%'),
        ('🛢️ 原油 ⭐⭐','华宝油气','3%→7%','布伦特85-90','布伦特<78','BTFP类重启',''),
        ('🥈 白银 ⭐⭐','盛达','4%→6%','现货48-52$','现货<45离场','金银比>80',''),
        ('🪨 煤炭 ⭐','神华','7%→9%','神华28-30元','神华<26','电厂库存天数',''),
        ('📜 国债','511260','14%→15%','速冻避险首选','10Y>3.5%加','',''),
    ]),
    # Node 5: Main Rally (green)
    ('节点五\n主升兑现','12月–2027Q1','美元放水·全资产主升\n收获期·分批兑现', GREEN_FILL, BOLD_GREEN, [
        ('💎 黄金','518880','兑现30%留40%','5500+兑现30%','5800+再20%','央行购金+QE','总仓69→45%'),
        ('⚙️ 铝','云铝','兑现40%留50%','23000+兑现40%','','全球PMI>50','⚠ 分批兑现'),
        ('🔶 铜','紫金','兑现30%留50%','85000+兑现30%','','LME库存降','不贪最后铜板'),
        ('🛢️ 原油','华宝油气','兑现30%留30%','120+兑现30%','140+再20%','OPEC+减产',''),
        ('🥈 白银','盛达','兑现30%留40%','70+兑现30%','','金银比<60',''),
        ('🪨 煤炭','神华','兑现30%留50%','40+兑现30%','','电厂日耗',''),
        ('📜 国债','511260','减至10%','收益率<3.0%减','','保留10%防御',''),
    ]),
]

row = 4
for node_label, timeframe, event, fill, title_font, assets in nodes:
    start_row = row
    for i, (asset, ticker, entry, pos, stop, signal, position_note) in enumerate(assets):
        r = row + i
        ws.cell(row=r, column=1, value=node_label if i==0 else '').font = title_font
        ws.cell(row=r, column=2, value=timeframe if i==0 else '')
        ws.cell(row=r, column=3, value=event if i==0 else '')
        ws.cell(row=r, column=4, value=asset).font = BOLD
        ws.cell(row=r, column=5, value=ticker).font = NORMAL
        ws.cell(row=r, column=6, value=entry).font = Font(size=9, color='2E7D32')
        ws.cell(row=r, column=7, value=pos).font = BOLD
        ws.cell(row=r, column=8, value=stop).font = Font(size=9, color='C62828')
        ws.cell(row=r, column=9, value=signal).font = Font(size=8, color='757575')
        ws.cell(row=r, column=10, value=position_note).font = Font(size=8, color='E65100', bold=True)
        for c in range(1, 11):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).alignment = AL
            ws.cell(row=r, column=c).border = thin_border
        ws.row_dimensions[r].height = 28
    # Merge node label / timeframe / event cells
    if len(assets) > 1:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row+len(assets)-1, end_column=1)
        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row+len(assets)-1, end_column=2)
        ws.merge_cells(start_row=start_row, start_column=3, end_row=start_row+len(assets)-1, end_column=3)
    row += len(assets)

# Total position summary row
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
ws.cell(row=row, column=1, value='📊 全周期仓位演化: 节点一 37%→50% | 节点二 50%→61% | 节点三 61%→53% | 节点四⭐ 53%→69% | 节点五 69%→45%').font = Font(bold=True, size=11, color='263238')
ws.cell(row=row, column=1).fill = TOTAL_FILL; ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
for c in range(1,11): ws.cell(row=row, column=c).border = thin_border
ws.row_dimensions[row].height = 28

# Red lines
row += 2
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
ws.cell(row=row, column=1, value='🚨 三道风控红线: ①单一标的≤总资产5% | ②商品期货保证金≤10% | ③10Y美债实际收益率>4.45% → 权益总仓降10%').font = Font(bold=True, size=10, color='C62828')
ws.cell(row=row, column=1).fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
for c in range(1,11): ws.cell(row=row, column=c).border = thin_border
ws.row_dimensions[row].height = 25

# Danger vars
row += 1
dangers = ['💣 破坏变量: ①美债遇冷(投标<2.0)→油价暴涨提前 | ②美国CPI>5%连3月→停战延后 | ③中国抛储→国内涨幅打折 | ④中东停火→速冻提前 | ⑤美联储提前降息→周期压缩']
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
ws.cell(row=row, column=1, value=dangers[0]).font = Font(size=9, color='757575')
ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
for c in range(1,11): ws.cell(row=row, column=c).border = thin_border

# Column widths
for c, w in enumerate([14,13,20,14,22,20,14,15,18,16], 1):
    ws.column_dimensions[get_column_letter(c)].width = w

# ═══════════════════════════ SHEET 2 ═══════════════════════════
ws2 = wb.create_sheet("动态调整")
ws2.merge_cells('A1:H1'); ws2['A1'] = '⚙️ 风险因素 × 仓位动态调整矩阵'
ws2['A1'].font = Font(bold=True, size=14, color='263238'); ws2['A1'].alignment = Alignment(horizontal='center'); ws2.row_dimensions[1].height = 30

# WTI mapping
ws2.merge_cells('A3:G3'); ws2['A3'] = 'WTI油价 × 仓位映射'
ws2['A3'].font = Font(bold=True, size=12, color='E65100')
wti_h = ['WTI','通胀等级','风险','资源仓位','黄金','债券/现金','操作']
for c, v in enumerate(wti_h, 1):
    cell = ws2.cell(row=4, column=c, value=v); cell.font = WHITE_FONT; cell.fill = HEAD_FILL; cell.alignment = AL_C; cell.border = thin_border
wti_d = [
    ['<$75','通缩','🔴 硬止损','≤10%','20%','70%','防御·现金为王'],
    ['$75-85','温和','🟡 观察','20-30%','20%','50%','预热·正常持仓'],
    ['$85-110','高通胀','🟢 进攻','40-50%','15%','35%','冲刺·加仓'],
    ['$110-150','狂奔','🟠 高波动','50-60%','10%','30%','冲刺·满仓'],
    ['>$150','恶性','🔴 极端','减至30%','25%','45%','预警·止盈'],
]
fills = [RED_FILL, PatternFill(start_color='FFF9C4',end_color='FFF9C4',fill_type='solid'), GREEN_FILL, ORANGE_FILL, RED_FILL]
for i, row in enumerate(wti_d):
    for j, v in enumerate(row):
        cell = ws2.cell(row=5+i, column=j+1, value=v); cell.font = NORMAL; cell.alignment = AL_C; cell.border = thin_border; cell.fill = fills[i]
    ws2.row_dimensions[5+i].height = 24

# Risk adjustment
ws2.merge_cells('A11:G11'); ws2['A11'] = '🔧 风险因素调整矩阵'
ws2['A11'].font = Font(bold=True, size=12, color='1565C0')
adj_h = ['因素','当前','若恶化','仓位','买点','止损','若改善','仓位']
for c, v in enumerate(adj_h, 1):
    ws2.cell(row=12, column=c, value=v).font = WHITE_FONT; ws2.cell(row=12, column=c).fill = HEAD_FILL; ws2.cell(row=12, column=c).alignment = AL_C; ws2.cell(row=12, column=c).border = thin_border
adj_d = [
    ['10Y美债','4.67%','>5.0%','-10%','等回调','-5%','<4.0%','+10%'],
    ['DXY','100.12','>102','-15%','不追高','全紧','<98','+10%'],
    ['WTI','$85.56','>$150','-20%','止盈','$130','<$70','重评'],
    ['VIX','17.09','>25','-15%','等恐慌','15%回撤','<15','维持'],
    ['美债拍卖','正常','流拍','-30%','全清','无条件','超额','+20%'],
]
for i, row in enumerate(adj_d):
    for j, v in enumerate(row):
        cell = ws2.cell(row=13+i, column=j+1, value=v); cell.font = NORMAL; cell.alignment = AL_C; cell.border = thin_border
    ws2.cell(row=13+i, column=3).font = BOLD_RED; ws2.cell(row=13+i, column=7).font = BOLD_GREEN
    ws2.row_dimensions[13+i].height = 22

# Total control reminder
ws2.merge_cells('A19:H19'); ws2['A19'] = '⚠️ 总仓位把控提醒'
ws2['A19'].font = Font(bold=True, size=12, color='C62828')
reminders = [
    '节点一(当前): 总仓37%→50%。预留50%现金应对速冻期。',
    '节点二: 总仓50%→61%。趋势确认加仓，不加杠杆。',
    '节点三: 总仓61%→53%。兑现利润是纪律，不是判断。',
    '节点四⭐: 总仓53%→69%。分三批接回，绝不在一天内满仓。',
    '节点五: 总仓69%→45%。分批兑现，保留40-50%底仓。',
    '全程: 单一标的≤5% | 期货≤10% | 10Y>4.45%降仓10%',
]
for i, r in enumerate(reminders):
    ws2.merge_cells(start_row=20+i, start_column=1, end_row=20+i, end_column=8)
    ws2.cell(row=20+i, column=1, value=r).font = Font(size=9, color='757575')
    ws2.cell(row=20+i, column=1).alignment = Alignment(vertical='center')

for c, w in enumerate([18,12,12,12,12,12,12,12], 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

out = r'c:\Users\Admin\Documents\陈嘉-资料备份\08.投资决策框架\看板日志\reports\special\向心坍缩推演\domestic_battle_map.xlsx'
wb.save(out)
print(f'SAVED: {out}')
